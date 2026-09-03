# -*- coding: utf-8 -*-
"""verification_dialog.py — 🧩 验证层 Feature/Test 节点 UI (2026-09-04 老倪)

三级结构 (节点 → 功能 → 用例): 真源 src/lerobot/verification/node_func_tree.py
  22 节点 × 5 功能 (名 5~10 字, 模块化) × 5 用例 (auto/semi/manual)
  纤维丛数学骨架: 功能分类 = 状态观测 / 工艺动作 / 闭环联络 / 质量截面

画布两个节点:
  🧩 Feature 功能清单 → 本对话框 (清单树 + 导出 Excel)
  🧪 Test 用例执行    → 本对话框 (跑套件 + 结果树 + 导出 Excel)
"""
import os

from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel,
                             QTreeWidget, QTreeWidgetItem, QPushButton,
                             QHeaderView, QAbstractItemView, QMessageBox)

_DARK = ("QDialog { background:#0d1117; color:#e6edf3; } "
         "QLabel { color:#e6edf3; } "
         "QTreeWidget { background:#161b22; color:#e6edf3; border:1px solid #30363d; "
         "outline:none; } "
         "QHeaderView::section { background:#21262d; color:#e6edf3; border:none; padding:4px; } "
         "QPushButton { background:#1f6feb; color:#fff; border:none; border-radius:5px; "
         "padding:8px 16px; font-size:13px; } "
         "QPushButton:hover { background:#388bfd; } "
         "QPushButton#b_export { background:#238636; } "
         "QPushButton#b_close { background:#30363d; }")

# 纤维丛功能分类着色 (数学骨架标签)
_FB_COLOR = {"底空间·状态观测": "#58a6ff", "纤维·工艺动作": "#3fb950",
             "闭环联络·状态预测": "#00d4aa", "闭环联络·决策调度": "#d29922",
             "闭环联络·状态校正": "#ff7b72", "闭环联络·安全": "#f85149",
             "闭环联络·工艺编排": "#a371f7", "质量截面·检测": "#ffd700",
             "质量截面·诊断": "#ffa657", "质量截面·几何监测": "#7ee787",
             "质量截面·对准评估": "#79c0ff", "底空间·低维结构(世界模型)": "#bc8cff",
             "底空间·状态转移": "#56d364", "底空间·状态观测(感知)": "#58a6ff"}


def _repo_root():
    d = os.path.dirname(os.path.abspath(__file__))
    while d != os.path.dirname(d):
        if os.path.isdir(os.path.join(d, "src", "lerobot")) and \
                os.path.isdir(os.path.join(d, "tools", "gui")):
            return d
        d = os.path.dirname(d)
    return d


def _load_tree():
    """加载三级树注册表 (node_func_tree.py)"""
    import importlib.util as _ilu
    path = os.path.join(_repo_root(), "src", "lerobot", "verification", "node_func_tree.py")
    spec = _ilu.spec_from_file_location("lerobot.verification.node_func_tree", path)
    m = _ilu.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


def export_verif_excel(path=None, tree=None, results=None):
    """三级树 → Excel (4 sheets: 节点功能清单/功能用例/分类统计/测试明细)
    tree: node_func_tree 模块; results: run_tree() 的 {case_key: (ok, detail)}"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    if tree is None:
        tree = _load_tree()
    if path is None:
        path = os.path.join(_repo_root(), "reports", "state_space_features.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = openpyxl.Workbook()
    _HDR = PatternFill("solid", fgColor="1F6FEB")
    _HF = Font(color="FFFFFF", bold=True, size=11)

    # ── Sheet1 功能清单 (规范场层 × 节点维度) ──
    ws = wb.active
    ws.title = "功能清单"
    ws.append(["规范场层", "节点", "节点名", "纤维丛分类", "功能ID", "功能名", "功能说明",
               "用例数", "自动", "半自动", "手动"])
    for c in ws[1]:
        c.fill, c.font = _HDR, _HF
    _GNAME = {g[0]: g[1] for g in tree.GAUGE_LAYERS}
    for nk, node in tree.NODE_TREE.items():
        g = tree.gauge_of(nk)
        for f in node["funcs"]:
            kinds = [t[1] for t in f["tests"]]
            ws.append([f"{g} {_GNAME.get(g, '')}", nk, node["name"], node["fb"],
                       f["fid"], f["name"], f["desc"],
                       len(f["tests"]), kinds.count("auto"), kinds.count("semi"),
                       kinds.count("manual")])
    for i, w in enumerate((16, 10, 26, 22, 10, 14, 40, 8, 8, 10, 8), start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Sheet2 功能用例 (每功能 5 用例明细) ──
    ws2 = wb.create_sheet("功能用例")
    ws2.append(["节点", "功能名", "用例#", "用例描述", "类型", "断言方法", "手动步骤"])
    for c in ws2[1]:
        c.fill, c.font = _HDR, _HF
    for nk, node in tree.NODE_TREE.items():
        for f in node["funcs"]:
            for ti, (desc, kind, ref, step) in enumerate(f["tests"]):
                ws2.append([nk, f["name"], ti + 1, desc, kind, ref or "", step or ""])
    for i, w in enumerate((10, 14, 8, 44, 8, 22, 40), start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── Sheet3 分类统计 ──
    ws3 = wb.create_sheet("分类统计")
    ws3.append(["维度", "类别", "数量"])
    for c in ws3[1]:
        c.fill, c.font = _HDR, _HF
    from collections import Counter
    # 规范场三层统计
    for gid, zh, en, d, nn, nf, nt in tree.gauge_stats():
        ws3.append(["规范场层", f"{zh} ({en})", f"{nn} 节点/{nf} 功能/{nt} 用例"])
    fb = Counter(n["fb"] for n in tree.NODE_TREE.values())
    for k, v in fb.items():
        ws3.append(["纤维丛分类", k, v])
    kc = tree.kind_count()
    for k, v in kc.items():
        ws3.append(["用例类型", k, v])
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 46

    # ── Sheet4 测试明细 (有结果时) ──
    if results:
        ws4 = wb.create_sheet("测试明细")
        ws4.append(["用例key", "节点", "功能", "结果", "证据"])
        for c in ws4[1]:
            c.fill, c.font = _HDR, _HF
        for key, (ok, detail) in sorted(results.items()):
            parts = key.split(".")
            node = parts[0] if parts else key
            ws4.append([key, node, parts[1] if len(parts) > 1 else "",
                        "PASS" if ok is True else ("FAIL" if ok is False else "SKIP/手动"),
                        str(detail)[:200]])
        for i, w in enumerate((24, 10, 12, 10, 80), start=1):
            ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(path)
    return path


class VerificationDialog(QDialog):
    """🧩 验证层 — 节点→功能→用例 三级树 (Feature 清单 / Test 结果)"""

    export_done = pyqtSignal(str)

    def __init__(self, mode="feature", parent=None, log=None):
        super().__init__(parent)
        self._mode = mode
        self._log = log or (lambda *a: None)
        self._tree = _load_tree()
        self._results = None
        self.setWindowTitle("🧩 功能清单 · 规范场三层 → 节点 → 功能 → 用例" if mode == "feature"
                            else "🧪 用例执行 · 规范场三层 → 节点 → 功能 → 用例")
        self.setStyleSheet(_DARK)
        self.setMinimumSize(1000, 560)
        self.resize(1280, 720)
        lay = QVBoxLayout(self)

        _n, _f = self._tree.node_count(), self._tree.func_count()
        _k = self._tree.kind_count()
        _gs = self._tree.gauge_stats()
        _g_txt = "  |  ".join(f"{zh} {t}用例" for gid, zh, en, d, nn, nf, t in _gs)
        self.lbl_sum = QLabel(
            f"规范场三层: {_g_txt}\n"
            f"{_n} 节点 × {_f} 功能 (名 5~10 字) × 5 用例 · 自动 {_k.get('auto', 0)} · "
            f"半自动 {_k.get('semi', 0)} · 手动 {_k.get('manual', 0)} · "
            f"模块化组合链 {len(self._tree.FUNC_CHAINS)} 条")
        self.lbl_sum.setStyleSheet("color:#8b949e; font-size:12px; padding:2px;")
        lay.addWidget(self.lbl_sum)

        # 树: 节点 → 功能 → 用例
        self.tr = QTreeWidget()
        self.tr.setStyleSheet(_DARK + "QTreeWidget::item { padding:2px; }")
        self.tr.setColumnCount(4)
        self.tr.setHeaderLabels(["名称", "分类", "类型", "说明/证据"])
        self.tr.setRootIsDecorated(True)
        self.tr.setAlternatingRowColors(True)
        self.tr.setEditTriggers(QAbstractItemView.NoEditTriggers)
        hdr = self.tr.header()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.Stretch)
        lay.addWidget(self.tr, 1)
        self._populate()

        h = QHBoxLayout()
        self.lbl_res = QLabel("")
        self.lbl_res.setStyleSheet("color:#8b949e; font-size:12px;")
        self.lbl_res.setTextInteractionFlags(Qt.TextSelectableByMouse)
        h.addWidget(self.lbl_res)
        h.addStretch()
        if mode == "feature":
            b_run = QPushButton("▶ 运行全部自动用例")
            b_run.setToolTip("后台跑三级树 auto 用例 (339 项, 引擎/六层/YOLO 真断言)")
            b_run.clicked.connect(self._run_tests)
            h.addWidget(b_run)
        b_export = QPushButton("导出 Excel")
        b_export.setObjectName("b_export")
        b_export.setToolTip("导出功能清单/用例 Excel → 上传 datadrive.world")
        b_export.clicked.connect(self._export)
        b_close = QPushButton("关闭")
        b_close.setObjectName("b_close")
        b_close.clicked.connect(self.close)
        h.addWidget(b_export)
        h.addWidget(b_close)
        lay.addLayout(h)
        self.export_done.connect(self.lbl_res.setText)
        self.lbl_res.setText("三级清单就绪 — ▶ 运行自动用例 或 导出 Excel")

    # ── 树填充 (规范场三层 → 节点 → 功能 → 用例) ──
    def _populate(self):
        self.tr.clear()
        # 三层分组: G1 场感知 / G2 协变操作 / G3 对称认知
        for gid, zh, en, desc, nks in self._tree.GAUGE_LAYERS:
            _nodes = [k for k in nks if k in self._tree.NODE_TREE]
            if not _nodes:
                continue
            g_item = QTreeWidgetItem([f"{zh}  {en}", gid, "",
                                      f"{len(_nodes)} 节点 · "
                                      f"{sum(len(self._tree.NODE_TREE[k]['funcs']) for k in _nodes)} 功能 · "
                                      f"{sum(len(fn['tests']) for k in _nodes for fn in self._tree.NODE_TREE[k]['funcs'])} 用例 · {desc}"])
            g_item.setForeground(0, Qt.yellow)
            self.tr.addTopLevelItem(g_item)
            for nk in _nodes:
                node = self._tree.NODE_TREE[nk]
                n_item = QTreeWidgetItem([f"  {node['name']}  ({nk})", node["fb"], "",
                                          f"{len(node['funcs'])} 功能 · {sum(len(f['tests']) for f in node['funcs'])} 用例"])
                g_item.addChild(n_item)
                for f in node["funcs"]:
                    kinds = [t[1] for t in f["tests"]]
                    f_item = QTreeWidgetItem([f"  ▪ {f['name']}  ({f['fid']})", "", "",
                                              f"{f['desc']} · auto {kinds.count('auto')}/semi {kinds.count('semi')}/手动 {kinds.count('manual')}"])
                    f_item.setForeground(0, Qt.cyan)
                    n_item.addChild(f_item)
                    for ti, (tdesc, kind, ref, step) in enumerate(f["tests"]):
                        _mark = {"auto": "●", "semi": "◐", "manual": "○"}.get(kind, "?")
                        _res = ""
                        if self._results is not None:
                            r = self._results.get(f"{nk}.{f['fid']}.{ti}")
                            if r:
                                _res = "✅ " if r[0] is True else ("❌ " if r[0] is False else "⏭ ")
                                _res += str(r[1])[:120]
                        t_item = QTreeWidgetItem([f"    {_mark} 用例{ti+1} · {tdesc}", "", kind,
                                                  _res or (step or ref or "")])
                        t_item.setForeground(2, {"auto": Qt.green, "semi": Qt.yellow,
                                                 "manual": Qt.gray}.get(kind))
                        f_item.addChild(t_item)
            g_item.setExpanded(True)

    # ── 后台跑测试 (不冻结 GUI) ──
    def _run_tests(self):
        self.lbl_res.setText("⏳ 后台运行自动用例 (引擎/六层/源码审计, YOLO 慢项跳过)…")
        import threading

        def _work():
            try:
                import importlib.util as _ilu
                vp = os.path.join(_repo_root(), "src", "lerobot", "verification",
                                  "verification_layer.py")
                spec = _ilu.spec_from_file_location("lerobot.verification.verification_layer", vp)
                m = _ilu.module_from_spec(spec)
                spec.loader.exec_module(m)
                v = m.VerificationLayer(log=lambda *a: None)
                ok, res = v.run_tree(skip_slow=True, log_fn=lambda *a: None)
                self._results = res
                passed = sum(1 for x in res.values() if x and x[0] is True)
                failed = sum(1 for x in res.values() if x and x[0] is False)
                from PyQt5.QtCore import QTimer
                QTimer.singleShot(0, lambda: (self._populate(),
                                              self.lbl_res.setText(
                                                  f"✅ 自动用例: PASS {passed} · FAIL {failed}"
                                                  f"{' · 全绿' if ok else ''}")))
            except Exception as e:
                self.export_done.emit(f"⚠️ 测试运行失败: {e}")

        threading.Thread(target=_work, daemon=True).start()

    # ── 导出 Excel (后台线程 + scp 上传) ──
    def _export(self):
        self.lbl_res.setText("⏳ 正在导出 Excel 并上传 datadrive.world…")
        import threading

        def _work():
            try:
                path = export_verif_excel(tree=self._tree, results=self._results)
                import subprocess as _sp
                fname = os.path.basename(path)
                msg = f"✅ 已导出: {path}"
                try:
                    r = _sp.run(["sshpass", "-p", "Nix19789", "scp", "-o", "StrictHostKeyChecking=no",
                                 "-o", "ConnectTimeout=15", path,
                                 f"root@39.102.211.79:/www/wwwroot/datadrive.world/{fname}"],
                                capture_output=True, timeout=60)
                    if r.returncode == 0:
                        _sp.run(["sshpass", "-p", "Nix19789", "ssh", "-o", "StrictHostKeyChecking=no",
                                 "-o", "ConnectTimeout=15", "root@39.102.211.79",
                                 f"chmod 644 /www/wwwroot/datadrive.world/{fname}"],
                                capture_output=True, timeout=30)
                        msg = (f'✅ 已导出并上传: <a href="http://datadrive.world/{fname}" '
                               f'style="color:#58a6ff;">http://datadrive.world/{fname}</a>')
                except Exception as _e:
                    msg += f" (上传失败: {_e})"
            except Exception as _e:
                msg = f"⚠️ 导出失败: {_e}"
            self.export_done.emit(msg)

        threading.Thread(target=_work, daemon=True).start()


if __name__ == "__main__":
    import sys as _s
    from PyQt5.QtWidgets import QApplication
    app = QApplication(_s.argv)
    d = VerificationDialog(mode=_s.argv[1] if len(_s.argv) > 1 else "feature")
    d.show()
    _s.exit(app.exec_())
