# -*- coding: utf-8 -*-
"""feature.dbc — 能力数据库 解析器/生成器 (2026-08-19 老倪)

参考 Vector CANoe DBC 设计: 同一平台/容器内用一份数据库文件配置不同模型。
- export_dbc() : model_feature.py 能力库 → feature.dbc 文本
- parse_dbc()  : feature.dbc 文本 → 结构化数据 (模型节点/能力/信号/组合)
- build_tree() : 从解析结果构建数据字典树 (文件即配置事实)
- 第三方模型接入: BU_ 加节点 + CM_ 声明能力 + SG_ 对齐信号 → 平台同流程运行
"""
import os
import re

# 仓库根 (tools/gui/../.. = 仓库根)
_REPO_ROOT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
DBC_PATH = os.path.join(_REPO_ROOT, "feature.dbc")


# ── 生成 ─────────────────────────────────────────────
def export_dbc(library, manifests, dataflow_stages, iface_defs):
    """能力库数据结构 → feature.dbc 文本"""
    L = ["VERSION \"1.0\""]
    L.append("// ═══════════════════════════════════════════════════")
    L.append("// feature.dbc — 能力数据库 (参考 Vector CANoe DBC)")
    L.append("// 同一平台/容器内配置不同模型的能力接入, 第三方模型按文件接入")
    L.append("// ═══════════════════════════════════════════════════")
    L.append("")
    # BU_: 模型节点
    L.append("// 模型节点 (平台内可配置的模型/第三方接入点)")
    nodes = list(manifests.keys())
    L.append("BU_: " + " ".join(n.upper() for n in nodes))
    L.append("")
    # FLOW_: 数据流形态
    L.append("// 数据流形态 (状态空间=形态之一)")
    for fname, fdesc in dataflow_stages:
        L.append(f"FLOW_: {fname.split()[0].upper()} \"{fname.split()[0]}\" \"{fdesc}\"")
    for key, m in manifests.items():
        L.append(f"FLOW_: {key.upper()} \"{m['dataflow']}\" \"{m['name']}\" {key.upper()}")
    L.append("")
    # BO_/SG_: 能力与信号
    L.append("// 能力 (BO_) 与 输入输出信号 (SG_), 接口取值: IN/OUT/CFG/TRAIN/DEPLOY/EVAL/MON/SCHED")
    for cat, items in library:
        L.append(f"// ── {cat} ──")
        for f in items:
            L.append(f"BO_ {f['id']} {f['name']}: {f['iface']} {cat.split()[0]}")
            L.append(f" SG_ 简述 : 64 \"{f['desc']}\"")
            L.append(f" SG_ 解释 : 512 \"{f['explain']}\"")
            L.append(f" SG_ 接口定义 : 512 \"{f['iface_def']}\"")
            L.append(f" SG_ 输入 : 24 \"{f['io_in']}\"")
            L.append(f" SG_ 输出 : 64 \"{f['io_out']}\"")
            L.append(f" SG_ 场景 : 128 \"{f['scene']}\"")
            L.append(f" SG_ 工程 : 128 \"{f['eng']}\"")
            L.append(f" SG_ 归属 : 64 \"{f['app']}\"")
    L.append("")
    # CM_: 模型能力组合
    L.append("// 模型能力组合 (节点: 选用能力)")
    for key, m in manifests.items():
        ids = " ".join(sorted(m["features"]))
        L.append(f"CM_ {key.upper()}: {ids}")
    return "\n".join(L) + "\n"


def write_dbc(library, manifests, dataflow_stages, iface_defs, path=DBC_PATH):
    """生成并写盘 feature.dbc"""
    txt = export_dbc(library, manifests, dataflow_stages, iface_defs)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(txt)
    return path


# ── 解析 ─────────────────────────────────────────────
def parse_dbc(text):
    """feature.dbc 文本 → dict {nodes, flows, capabilities, manifests}
    capabilities: {id: {name, iface, cat, io_in, io_out}}
    manifests:    {node: set(feature ids)}
    """
    out = {"nodes": [], "flows": [], "capabilities": {}, "manifests": {}}
    cur_cap = None
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("//"):
            continue
        if line.startswith("VERSION"):
            out["version"] = line.split('"')[1] if '"' in line else ""
        elif line.startswith("BU_:"):
            out["nodes"] = line.split(":", 1)[1].split()
        elif line.startswith("FLOW_:"):
            out["flows"].append(line[6:].strip())
        elif line.startswith("BO_ "):
            m = re.match(r"BO_\s+(\S+)\s+(.+?):\s+(\S+)\s+(\S+)", line)
            if m:
                fid, name, iface, cat = m.group(1), m.group(2), m.group(3), m.group(4)
                cur_cap = fid
                out["capabilities"][fid] = {"id": fid, "name": name, "iface": iface,
                                            "cat": cat, "desc": "", "explain": "",
                                            "iface_def": "", "io_in": "", "io_out": "",
                                            "scene": "", "eng": "", "app": ""}
        elif line.startswith("SG_") and cur_cap:
            m = re.match(r'SG_\s+(\S+)\s*:\s*\d+\s+"(.*)"', line)
            if m:
                sname, sdesc = m.group(1), m.group(2)
                key = {"简述": "desc", "解释": "explain", "接口定义": "iface_def",
                       "输入": "io_in", "输出": "io_out", "场景": "scene",
                       "工程": "eng", "归属": "app"}.get(sname)
                if key:
                    out["capabilities"][cur_cap][key] = sdesc
        elif line.startswith("CM_"):
            m = re.match(r"CM_\s+(\S+):\s*(.*)", line)
            if m:
                out["manifests"][m.group(1).upper()] = set(m.group(2).split())
    return out


def load_dbc(path=DBC_PATH):
    """读文件解析; 文件不存在返回 None"""
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as fh:
        return parse_dbc(fh.read())


# ── 树构建 (文件即配置事实) ──────────────────────────
def export_excel(path=None, dbc=None, library=None, manifests=None, iface_defs=None):
    """能力库 → Excel (3 sheets: 能力库/模型组合/接口说明)
    path=None → reports/feature_dbc.xlsx"""
    if dbc is None:
        dbc = load_dbc()
    if library is None:
        try:
            import model_feature as _mf
            library, manifests, iface_defs = (_mf.FEATURE_LIBRARY,
                                              _mf.MODEL_MANIFESTS,
                                              _mf.INTERFACE_DEFS)
        except Exception:
            library, manifests, iface_defs = [], {}, {}
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    if path is None:
        path = os.path.join(_REPO_ROOT, "reports", "feature_dbc.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = openpyxl.Workbook()

    _HDR = PatternFill("solid", fgColor="1F6FEB")
    _HF = Font(color="FFFFFF", bold=True, size=11)

    # ── Sheet1 能力库 ──
    ws = wb.active
    ws.title = "能力库"
    cols = ["ID", "能力", "简述", "解释说明", "接口定义", "输入信号", "输出信号",
            "接口", "场景", "工程落点", "归属"]
    ws.append(cols)
    for c in ws[1]:
        c.fill, c.font = _HDR, _HF
    for cat, items in library:
        for f in items:
            ws.append([f["id"], f["name"], f["desc"], f["explain"], f["iface_def"],
                       f["io_in"], f["io_out"], f["iface"], f["scene"],
                       f["eng"], f["app"]])
    for i, w in enumerate((6, 14, 26, 52, 52, 42, 42, 12, 26, 24, 22), start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Sheet2 模型组合 ──
    ws2 = wb.create_sheet("模型组合")
    ws2.append(["模型节点", "数据流形态", "能力数", "能力清单"])
    for c in ws2[1]:
        c.fill, c.font = _HDR, _HF
    for key, m in manifests.items():
        ws2.append([key.upper(), m.get("dataflow", ""), len(m["features"]),
                    " ".join(sorted(m["features"]))])
    for i, w in enumerate((16, 40, 10, 80), start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Sheet3 接口说明 ──
    ws3 = wb.create_sheet("接口说明")
    ws3.append(["接口", "说明"])
    for c in ws3[1]:
        c.fill, c.font = _HDR, _HF
    for k, v in (iface_defs or {}).items():
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 60

    wb.save(path)
    return path


def upload_excel(path=None):
    """导出 Excel 并上传 datadrive.world (ECS), 返回 (本地路径, 下载URL)
    用户无 /mnt/c 共享 → 通过网站下载 (记忆链路: sshpass scp + chmod644)"""
    import subprocess
    path = path or export_excel()
    url = "http://datadrive.world/feature_dbc.xlsx"
    try:
        r = subprocess.run(
            ["sshpass", "-p", "Nix19789", "scp", "-o", "StrictHostKeyChecking=no",
             "-o", "ConnectTimeout=15", path,
             "root@39.102.211.79:/www/wwwroot/datadrive.world/feature_dbc.xlsx"],
            capture_output=True, timeout=60)
        if r.returncode != 0:
            return path, None
        subprocess.run(
            ["sshpass", "-p", "Nix19789", "ssh", "-o", "StrictHostKeyChecking=no",
             "-o", "ConnectTimeout=15", "root@39.102.211.79",
             "chmod 644 /www/wwwroot/datadrive.world/feature_dbc.xlsx"],
            capture_output=True, timeout=60)
        return path, url
    except Exception:
        return path, None


def build_tree_from_dbc(dbc, module, make_item, user_role):
    """从解析的 dbc 构建能力库树 (供 model_tree.py 使用)
    make_item(texts) → QTreeWidgetItem; user_role 用于 setData"""
    root = make_item(["🧩 能力数据库 feature.dbc · 精细操作", f"v{dbc.get('version', '1.0')}"])
    root.setData(0, user_role, None)

    # 数据流形态
    df = make_item(["🔄 数据流形态 (状态空间=形态之一)", ""])
    df.setData(0, user_role, None)
    root.addChild(df)
    for fl in dbc.get("flows", []):
        # FLOW_ 行: ID "名称" "说明" [节点...]
        parts = re.findall(r'"([^"]*)"|\S+', fl)
        if len(parts) >= 2:
            fid = parts[0]
            name = parts[1] if '"' in fl else fl
            desc = parts[2] if len(parts) > 2 else ""
            _it = make_item([f"{fid} {name}", desc])
            df.addChild(_it)

    # 能力 (按库顺序: 解析时按 BO_ 出现顺序)
    caps = dbc.get("capabilities", {})
    # 按大类聚合 (从 capability.cat)
    cats = {}
    for fid, c in caps.items():
        cats.setdefault(c["cat"], []).append(c)
    sel = set()
    cur_node = None
    # 当前模型节点 (从 module 推断)
    try:
        from model_feature import current_model_key
        key = current_model_key(module)
        if key:
            cur_node = key.upper()
            sel = dbc.get("manifests", {}).get(cur_node, set())
    except Exception:
        pass

    for cat, items in cats.items():
        cn = make_item([f"📂 {cat} ({len(items)})", ""])
        cn.setData(0, user_role, None)
        root.addChild(cn)
        for c in items:
            fid = c["id"]
            mark = "✓ " if fid in sel else "○ "
            ft = make_item([f"{mark}{fid} {c['name']}", c.get("desc", "")])
            ft.setData(0, user_role, None)
            cn.addChild(ft)
            for _label, _val in (("解释说明", c.get("explain", "")),
                                ("接口定义", c.get("iface_def", "")),
                                ("输入信号", c.get("io_in", "")),
                                ("输出信号", c.get("io_out", "")),
                                ("接口", c["iface"]),
                                ("场景", c.get("scene", "")),
                                ("工程", c.get("eng", "")),
                                ("归属", c.get("app", ""))):
                _it = make_item([_label, _val])
                ft.addChild(_it)

    # 模型组合
    mn = make_item(["📦 模型节点 (feature.dbc 配置)", ""])
    mn.setData(0, user_role, None)
    root.addChild(mn)
    for node, ids in dbc.get("manifests", {}).items():
        cur = " ← 当前" if node == cur_node else ""
        mi = make_item([f"{node}{cur}", " ".join(sorted(ids))])
        mi.setData(0, user_role, None)
        mn.addChild(mi)
    return root


if __name__ == "__main__":
    # 独立运行: 从 model_feature.py 生成 feature.dbc
    import model_feature as mf
    path = write_dbc(mf.FEATURE_LIBRARY, mf.MODEL_MANIFESTS,
                     mf.DATAFLOW_STAGES, mf.INTERFACE_DEFS)
    print(f"✅ 已生成: {path}")
    dbc = load_dbc(path)
    print(f"   节点: {dbc['nodes']}")
    print(f"   能力: {len(dbc['capabilities'])} 条")
    print(f"   组合: {list(dbc['manifests'].keys())}")
