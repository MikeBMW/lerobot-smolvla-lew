# -*- coding: utf-8 -*-
"""planner.py — 大模型层 · 云端任务规划 (状态空间模型画布)

慢决策层, 完全在实时控制回路之外:
  🧠 TaskPlanner       任务规划器: 自然语言/MES 工单指令 → 技能Token序列 → 下发状态机
  🔍 ExceptionReasoner 异常推理器: 阶段卡死/连续否决/未接触 → 异常诊断 + 恢复建议
  🛠  SkillComposer    技能编排器: 新型号规格 → 新技能序列 + 关键参数 (力阈值/节拍)

设计原则 (老倪 2026-08-20):
- 大模型管"想", 小模型管"动" — 规划只在任务开始时调用一次 (~1-2s), 不进实时回路
- LLM 可插拔: 配置 llm_url/llm_key 时走大模型 API; 未配置则规则回退 (确定性优先)
- 输出经过规则校验 (如"插入必须在取料之后"), 无效序列被拒绝 → 状态机只收合法 Token
- 零重依赖 (无 torch): 仅 json/os, 与六层源码同目录加载

技能库: flows/atomic_skill_tokens.json (242 条原子技能, 9 大类别: NPO/XPO/操作动作/…)
Token 格式: [SKILL_xxx] (如 [SKILL_NPO015] 低力取出)
"""
import json
import os

_SKILL_TOKENS_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))),
    "flows", "atomic_skill_tokens.json")

# ── 指令 → 技能阶段的规则关键词 (规则回退用, 顺序=优先级) ────────────────
# 唯一映射源: _pick_skill/_pick_for_stage 与 _stage_of 共用, 保证可逆
# 注意: 阶段名本身放首位 (选择质量最高), 单字"插"覆盖 插电口/插光口/插光纤
_KW_MAP = {
    "取料": ["取料", "抓取", "取盘", "取出", "取放", "拿起", "拾取", "来料", "取"],
    "扫码": ["扫码", "条码", "二维码", "OCR", "读取", "序列号"],
    "对准": ["对准", "对接", "对位", "同轴", "定位", "姿态", "入位"],
    "插入": ["插入", "插接", "插电口", "插光口", "插光纤", "压接", "装入", "接合", "锁止", "受控推进"],
    "检测": ["检测", "复测", "测试", "校验", "确认", "判断", "识别", "飞拍", "目检"],
    "拔出": ["拔出", "拉出", "拔", "拆卸", "抬升", "脱离", "释放"],
    "分拣": ["分拣", "下料", "放料", "放置", "回装", "移载"],
    "运输": ["运输", "配送", "导航", "就位", "搬运", "转运"],
}
_STAGE_ORDER = ["取料", "运输", "扫码", "对准", "插入", "检测", "拔出", "分拣"]

# 新型号规格关键词 → 技能族 (技能编排器规则回退用)
_MODEL_KEYWORDS = {
    "QSFP":   "XPO高密可插拔光学",
    "OSFP":   "XPO高密可插拔光学",
    "CFP":    "XPO高密可插拔光学",
    "NPO":    "NPO近封装光学",
    "光电引擎": "NPO近封装光学",
    "光模块":  "XPO高密可插拔光学",
}

# 型号 → 建议参数 (力阈值 N / 节拍 s) — 规则回退基线, 真机标定后覆盖
_MODEL_PARAMS = {
    "QSFP":  {"force_limit": 20.0, "tact_time": 15.0, "insert_depth": 0.004},
    "OSFP":  {"force_limit": 25.0, "tact_time": 18.0, "insert_depth": 0.004},
    "CFP":   {"force_limit": 30.0, "tact_time": 20.0, "insert_depth": 0.005},
    "NPO":   {"force_limit": 8.0,  "tact_time": 25.0, "insert_depth": 0.002},
    "光电引擎": {"force_limit": 8.0, "tact_time": 25.0, "insert_depth": 0.002},
    "光模块":  {"force_limit": 20.0, "tact_time": 15.0, "insert_depth": 0.004},
}


def _load_skill_library():
    """加载技能库 → {skill_id: skill} + 关键词索引 {name: skill_id}"""
    if not os.path.exists(_SKILL_TOKENS_PATH):
        return {}, {}
    with open(_SKILL_TOKENS_PATH, encoding="utf-8") as f:
        data = json.load(f)
    skills = {s["skill_id"]: s for s in data.get("skills", [])}
    # name → skill_id (含类别前缀, 命中更准)
    by_name = {}
    for s in skills.values():
        by_name[s["name"]] = s["skill_id"]
        by_name[f"{s['category']}{s['name']}"] = s["skill_id"]
    return skills, by_name


# 五大作业场景权威数据源 (2026-08-20): flows/scenes_5jobs.json
_SCENES_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))),
    "flows", "scenes_5jobs.json")


def _load_scenes():
    """加载五大作业场景 → (scenes 列表, {scene_id: scene})"""
    if not os.path.exists(_SCENES_PATH):
        return [], {}
    with open(_SCENES_PATH, encoding="utf-8") as f:
        data = json.load(f)
    scenes = data.get("scenes", [])
    return scenes, {s["scene_id"]: s for s in scenes}


class TaskPlanner:
    """🧠 任务规划器 — 指令 → 技能Token序列 (慢决策, 回路外)

    输入: 自然语言指令 / MES 工单 (如 "先插QSFP模块, 测完把Pass的放到右边料盘")
    输出: 技能Token序列 [SKILL_xxx, ...] → 下发 S3 状态机执行
    模式: llm_url 配置 → 大模型拆解; 否则规则拆解 (确定性, 默认)
    """

    def __init__(self, llm_url=None, llm_key=None, llm_model="Qwen3-7B"):
        self.llm_url = llm_url
        self.llm_key = llm_key
        self.llm_model = llm_model
        self.skills, self.skill_by_name = _load_skill_library()

    # ── 对外接口 ──
    def plan(self, instruction):
        """指令 → 合法技能Token序列 (校验后)"""
        if not instruction or not instruction.strip():
            instruction = "插入光模块"
        tokens = self._plan_llm(instruction) if self.llm_url else self._plan_rules(instruction)
        tokens = self.validate(tokens)
        return tokens

    # ── 规则拆解 (确定性优先) ──
    def _plan_rules(self, instruction):
        """关键词 → 阶段顺序 → 技能库逐阶段选技能 (阶段不足 → 完整主链)"""
        stages = [st for st, kws in _KW_MAP.items()
                  if any(k in instruction for k in kws)]
        if len(stages) < 2:
            stages = ["取料", "对准", "插入", "检测"]   # 默认插拔主链
        # 按主链顺序排列 (指令可乱序描述, 执行必须有序)
        stages = [s for s in _STAGE_ORDER if s in stages] or stages
        return [self._pick_skill(stage) for stage in stages]

    def _pick_skill(self, stage):
        """阶段 → 技能库中匹配的技能 ID (kw 顺序=优先级, 与 _stage_of 同源)"""
        kws = _KW_MAP.get(stage, [])
        for kw in kws:
            for s in self.skills.values():
                if kw in s["name"]:
                    return s["tokens"]["id"]
        # 兜底: 类别泛化技能
        for s in self.skills.values():
            if s["category"] == "操作动作" and s["action"] == "operate":
                return s["tokens"]["id"]
        return "[SKILL_UNKNOWN]"

    # ── 大模型拆解 (可插拔) ──
    def _plan_llm(self, instruction):
        """调 LLM 拆解 → 技能ID序列 (失败回退规则)"""
        try:
            import urllib.request
            prompt = (
                "你是Z-MAX光模块工厂的机器人任务规划器。将指令拆解为原子技能Token序列, "
                "只输出技能ID列表(JSON数组)。可用技能ID示例: " +
                ", ".join(list(self.skill_by_name.values())[:40]) +
                f"。指令: {instruction}")
            req = urllib.request.Request(
                self.llm_url,
                data=json.dumps({"model": self.llm_model,
                                 "messages": [{"role": "user", "content": prompt}],
                                 "temperature": 0.1}).encode(),
                headers={"Content-Type": "application/json",
                         "Authorization": f"Bearer {self.llm_key or ''}"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                text = json.loads(resp.read().decode())["choices"][0]["message"]["content"]
            ids = [t.strip() for t in text.replace("[", "").replace("]", "")
                   .replace('"', "").split(",") if t.strip().startswith("SKILL_")]
            if ids:
                return ids
        except Exception:
            pass
        return self._plan_rules(instruction)

    # ── 规则校验 (非法序列拒绝) ──
    def validate(self, tokens):
        """校验: ①Token 存在于技能库 ②按主链阶段顺序排序 (LLM 乱序→归位)"""
        valid_ids = {s["tokens"]["id"] for s in self.skills.values()}
        valid = [t for t in tokens if t in valid_ids]
        if not valid:
            valid = [self._pick_skill(s) for s in _STAGE_ORDER]
        # 按阶段主链排序 (规则链天然有序; LLM 输出乱序时归位)
        order_map = {st: i for i, st in enumerate(_STAGE_ORDER)}
        ranked = []
        for t in valid:
            stage = self._stage_of(t)
            ranked.append((order_map.get(stage, len(_STAGE_ORDER)), t))
        ranked.sort(key=lambda x: x[0])
        return [t for _, t in ranked]

    def _stage_of(self, token):
        """Token → 阶段 (与 _pick_skill 同源反推, 取主链顺序最靠前的匹配)"""
        for s in self.skills.values():
            if s["tokens"]["id"] != token:
                continue
            for stage in _STAGE_ORDER:
                if any(k in s["name"] for k in _KW_MAP[stage]):
                    return stage
            return None
        return None


class ExceptionReasoner:
    """🔍 异常推理器 — 状态机卡住时诊断异常 + 恢复建议 (慢决策, 回路外)

    触发: S3 调度器连续否决 (max_veto) / 阶段停留超时 / 接触概率过低
    输入: 当前阶段 / 残差 / 接触概率 / 距离 / 停留时长
    输出: {type: 异常分类, advice: 恢复建议}
    """

    def __init__(self, llm_url=None, llm_key=None, llm_model="Qwen3-7B"):
        self.llm_url = llm_url
        self.llm_key = llm_key
        self.llm_model = llm_model

    def diagnose(self, stage="接近", residual=0.0, contact_p=0.5,
                 dist_h=0.05, dwell_time=0.0, veto_count=0, max_veto=3):
        """规则诊断 (LLM 未配置时) — 返回 (异常类型, 恢复建议)"""
        # ① 连续否决 → 力/接触异常
        if veto_count >= max_veto:
            return ("力控异常", f"连续 {veto_count} 次否决 (残差 {residual:.3f}) — 建议: 减速重试 + 复核力阈值")
        # ② 接近阶段卡死 (停留超时未接触)
        if stage == "接近" and dwell_time > 5.0 and contact_p < 0.6:
            return ("对准失败", f"接近阶段停留 {dwell_time:.1f}s 未接触 — 建议: 视觉复核孔位坐标 + 重新对准")
        # ③ 插入阶段未到位
        if stage == "插入" and contact_p > 0.6 and dist_h > 0.004:
            return ("插入未到位", f"距离 {dist_h*1000:.1f}mm 超插入阈值 — 建议: 复测 + 低力重插")
        # ④ 接触概率异常低
        if contact_p < 0.3:
            return ("未接触", "接触概率过低 — 建议: 检查末端位置与目标坐标")
        return (None, "运行正常")

    def diagnose_llm(self, context: dict):
        """LLM 诊断 (可插拔, 失败回退规则)"""
        try:
            import urllib.request
            prompt = ("Z-MAX光模块插拔异常诊断。当前阶段: {stage}; 残差: {residual}; "
                      "接触概率: {contact_p}; 距离: {dist_h}; 否决次数: {veto_count}。"
                      "输出JSON: {{\"type\": 异常分类, \"advice\": 恢复建议}}").format(**context)
            req = urllib.request.Request(
                self.llm_url,
                data=json.dumps({"model": self.llm_model,
                                 "messages": [{"role": "user", "content": prompt}],
                                 "temperature": 0.0}).encode(),
                headers={"Content-Type": "application/json",
                         "Authorization": f"Bearer {self.llm_key or ''}"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                return json.loads(resp.read().decode())["choices"][0]["message"]["content"]
        except Exception:
            return self.diagnose(**context)


class SkillComposer:
    """🛠 技能编排器 — 场景定义/新型号 → 新技能序列 + 关键参数 (新场景训练用)

    输入 (2026-08-20 升级, 场景驱动):
      a) 场景JSON (flows/scenes_5jobs.json 的 scene 对象) — 五大作业场景
         → 解析 objects/steps/performance: 步骤→技能序列, 性能约束→参数覆盖默认
      b) scene_id 字符串 ("SCN-01-FW") — 从场景库查
      c) 规格文本 (旧版兼容) — 关键词识别型号, 查 _MODEL_PARAMS 规则表
    输出: {scene_id, sequence: [SKILL_xxx,...], params: {force_limit, tact_time, ...}, targets}
    """

    def __init__(self, llm_url=None, llm_key=None, llm_model="Qwen3-7B"):
        self.llm_url = llm_url
        self.llm_key = llm_key
        self.llm_model = llm_model
        self.skills, _ = _load_skill_library()
        self.scenes, self.scene_by_id = _load_scenes()

    def compose(self, spec):
        """场景定义 / scene_id / 规格文本 → {sequence, params}"""
        scene = None
        if isinstance(spec, dict):
            scene = spec
        elif isinstance(spec, str):
            sid = spec.strip().upper()
            if sid in self.scene_by_id:
                scene = self.scene_by_id[sid]
        if scene is not None:
            return self._compose_from_scene(scene)
        return self._compose_from_spec(spec)

    # ── 场景驱动编排 (五大作业场景) ──
    def _compose_from_scene(self, scene):
        """场景JSON → 步骤(工艺) → 技能序列 + performance → 参数"""
        # 型号偏好: objects 里的插座/模块名 → 技能类别优先
        obj_text = " ".join(o.get("name", "") + o.get("role", "") for o in scene.get("objects", []))
        model_cat = None
        for kw, cat in _MODEL_KEYWORDS.items():
            if kw in obj_text or kw in scene.get("name", ""):
                model_cat = cat
                break
        # 步骤 → 阶段序列 (name+desc 拼接匹配, 覆盖"批量插入: 取模块→插入"这类)
        stages = []
        for st in scene.get("steps", []):
            text = st.get("name", "") + " " + st.get("desc", "")
            for stage in _STAGE_ORDER:
                if any(k in text for k in _KW_MAP[stage]):
                    if not stages or stages[-1] != stage:
                        stages.append(stage)
        if not stages:
            stages = ["取料", "对准", "插入", "检测"]
        # 阶段 → 技能 (型号类别优先, 兜底全库)
        sequence = []
        for stage in stages:
            tok = self._pick_for_stage(stage, model_cat)
            if tok:
                sequence.append(tok)
        # performance 约束 → 参数 (覆盖规则默认, 场景定义说了算)
        perf = scene.get("performance", {})
        params = {
            "force_limit": float(perf.get("force_limit", 20.0)),
            "tact_time": float(perf.get("tact_time", 15.0)),
            "positioning_accuracy": float(perf.get("positioning_accuracy", 0.0005)),
            "insert_depth": float(perf.get("insert_depth", 0.004)),
        }
        params.update({k: v for k, v in perf.items() if k not in params})
        params["scene_id"] = scene.get("scene_id")
        return {"scene_id": scene.get("scene_id"), "sequence": sequence,
                "params": params, "targets": scene.get("targets", {})}

    def _pick_for_stage(self, stage, model_cat=None):
        """阶段 → 技能ID: 型号类别只在阶段名上优先, 其余按全库精确匹配"""
        kws = _KW_MAP.get(stage, [])
        stage_kw = kws[0] if kws else stage
        if model_cat:
            for s in self.skills.values():
                if s["category"] == model_cat and stage_kw in s["name"]:
                    return s["tokens"]["id"]
        for kw in kws:
            for s in self.skills.values():
                if kw in s["name"]:
                    return s["tokens"]["id"]
        return None

    # ── 规格文本编排 (旧版兼容, 型号关键词) ──
    def _compose_from_spec(self, spec):
        """规格 → 技能序列 + 参数 (规则回退)"""
        spec = spec or ""
        model = None
        for kw, cat in _MODEL_KEYWORDS.items():
            if kw in spec:
                model = kw
                break
        seq = [self._pick(cat, ["取料", "取出", "取放", "抓取"]),
               self._pick(cat, ["对准", "定位", "同轴"]),
               self._pick(cat, ["插入", "插接", "压接", "装入", "锁止"]),
               self._pick(cat, ["检测", "确认", "校验"]),
               self._pick(cat, ["分拣", "放置", "回装"])]
        seq = [t for t in seq if t]
        params = dict(_MODEL_PARAMS.get(model, {"force_limit": 20.0, "tact_time": 15.0,
                                                "insert_depth": 0.004}))
        params["model"] = model or "未知型号"
        return {"sequence": seq, "params": params}

    def _pick(self, category, kws):
        for s in self.skills.values():
            if s["category"] == category and any(k in s["name"] for k in kws):
                return s["tokens"]["id"]
        for s in self.skills.values():
            if any(k in s["name"] for k in kws):
                return s["tokens"]["id"]
        return None

    # ── 全部任务导出 (2026-08-20 老倪: Excel 导出按钮) ──
    def export_all_tasks(self, path=None):
        """编排全部场景 → Excel (任务总览 + 技能明细 + 场景步骤 三个 sheet)

        Args:
            path: 输出 xlsx 路径; None → 默认 reports/task_export.xlsx
        Returns:
            (输出路径, 任务数) — 失败抛异常 (日志区显示根因)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        if path is None:
            path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
                    os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))),
                "reports", "task_export.xlsx")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="4472C4")
        # ── Sheet1 任务总览 ──
        ws = wb.active
        ws.title = "任务总览"
        headers = ["场景ID", "场景名称", "工艺目标", "技能Token序列", "力限N",
                   "节拍s", "对位精度m", "插入深度m", "额外约束"]
        ws.append(headers)
        for sid, scene in self.scene_by_id.items():
            out = self._compose_from_scene(scene)
            pr = out["params"]
            extra = {k: v for k, v in pr.items() if k not in
                     ("force_limit", "tact_time", "positioning_accuracy",
                      "insert_depth", "scene_id", "model")}
            ws.append([
                sid, scene.get("name", ""),
                " / ".join(f"{k}{v}" for k, v in out.get("targets", {}).items()),
                " → ".join(out["sequence"]),
                pr.get("force_limit"), pr.get("tact_time"),
                pr.get("positioning_accuracy"), pr.get("insert_depth"),
                json.dumps(extra, ensure_ascii=False),
            ])
        for c in ws[1]:
            c.font, c.fill = hdr_font, hdr_fill
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="center", wrap_text=True)
        widths = [14, 30, 22, 60, 8, 8, 14, 12, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        # ── Sheet2 技能明细 (每任务每步技能一行) ──
        ws2 = wb.create_sheet("技能明细")
        ws2.append(["场景ID", "步骤", "技能Token", "技能名称", "技能类别", "阶段"])
        stage_names = {i: s for i, s in enumerate(_STAGE_ORDER)}
        for sid, scene in self.scene_by_id.items():
            out = self._compose_from_scene(scene)
            for i, tok in enumerate(out["sequence"], 1):
                skill = None
                for s in self.skills.values():
                    if s["tokens"]["id"] == tok:
                        skill = s
                        break
                ws2.append([sid, i, tok,
                            skill["name"] if skill else "",
                            skill["category"] if skill else "",
                            ""])
        for c in ws2[1]:
            c.font, c.fill = hdr_font, hdr_fill
        for i, w in enumerate([14, 8, 20, 44, 18, 10], 1):
            ws2.column_dimensions[chr(64 + i)].width = w
        # ── Sheet3 场景步骤 (原始工艺步骤 + 力控约束) ──
        ws3 = wb.create_sheet("场景步骤")
        ws3.append(["场景ID", "步骤", "时间T", "时长s", "步骤名", "操作描述", "力控/精度约束"])
        for sid, scene in self.scene_by_id.items():
            for i, st in enumerate(scene.get("steps", []), 1):
                cons = " ".join(f"{k}={v}" for k, v in st.items()
                                if k in ("force", "accuracy", "depth", "count", "optional", "async"))
                ws3.append([sid, i, st.get("t"), st.get("dur"),
                            st.get("name"), st.get("desc"), cons])
        for c in ws3[1]:
            c.font, c.fill = hdr_font, hdr_fill
        for i, w in enumerate([14, 8, 10, 8, 18, 70, 30], 1):
            ws3.column_dimensions[chr(64 + i)].width = w
        wb.save(path)
        return path, len(self.scene_by_id)
