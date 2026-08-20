# -*- coding: utf-8 -*-
"""detection_targets.py — 🎯 YOLO 目标检测: 检测目标清单 + 评价指标 + Excel 导出

数据源: flows/detection_targets.json (2026-08-20 老倪, 来源=五大作业场景需求说明书)
六类 22 个检测目标: 类别识别 / 目标检测2D / 位姿估计 / 扫码OCR / 缺陷检测AOI / 状态识别
每条含: 检出对象、位姿/方向需求、评价指标 (mAP@0.5 / mAP@0.5:0.95 / 准确率 / 误差)、推理时间
"""
import json
import os

_DET_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
        os.path.dirname(os.path.abspath(__file__)))))),
    "flows", "detection_targets.json")


def load_detection_targets():
    """加载检测目标清单 → dict (targets/metrics_def/model_baseline)"""
    if not os.path.exists(_DET_PATH):
        return {"targets": [], "metrics_def": [], "model_baseline": []}
    with open(_DET_PATH, encoding="utf-8") as f:
        return json.load(f)


def export_excel(path=None):
    """检测目标清单 → Excel (检测目标清单 + 指标定义 + 模型基线 三 sheet)

    Args:
        path: 输出 xlsx; None → reports/detection_targets.xlsx
    Returns:
        (输出路径, 目标数)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    if path is None:
        path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
                os.path.dirname(os.path.abspath(__file__)))))),
            "reports", "detection_targets.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    data = load_detection_targets()
    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5597")
    # ── Sheet1 检测目标清单 ──
    ws = wb.active
    ws.title = "检测目标清单"
    ws.append(["目标ID", "检测目标", "类别", "检出对象", "检测内容",
               "来源场景", "位姿需求", "方向判定", "评价指标", "推理时间"])
    for t in data.get("targets", []):
        m = t.get("metrics", {})
        infer = m.pop("推理时间", None)
        metric_str = " / ".join(f"{k} {v}" for k, v in m.items())
        ws.append([t.get("target_id"), t.get("target"), t.get("category"),
                   "、".join(t.get("objects", [])), t.get("desc"),
                   "、".join(t.get("scenes", [])), t.get("pose"),
                   t.get("direction", "—"), metric_str, infer or ""])
    for c in ws[1]:
        c.font, c.fill = hdr_font, hdr_fill
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate([12, 20, 12, 34, 46, 30, 18, 18, 44, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    # ── Sheet2 指标定义 ──
    ws2 = wb.create_sheet("指标定义")
    ws2.append(["指标", "定义", "计算方式"])
    for d in data.get("metrics_def", []):
        ws2.append([d.get("metric"), d.get("def"), d.get("formula")])
    for c in ws2[1]:
        c.font, c.fill = hdr_font, hdr_fill
    for i, w in enumerate([20, 55, 50], 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    # ── Sheet3 模型基线 ──
    ws3 = wb.create_sheet("模型基线")
    ws3.append(["模型变体", "输入尺寸", "mAP@0.5", "mAP@0.5:0.95",
                "4090推理ms", "Orin推理ms", "参数量M", "备注"])
    for b in data.get("model_baseline", []):
        ws3.append([b.get("variant"), b.get("input"), b.get("mAP@0.5"),
                    b.get("mAP@0.5:0.95"), b.get("infer_4090_ms"),
                    b.get("infer_orin_ms"), b.get("params_M"), b.get("note")])
    for c in ws3[1]:
        c.font, c.fill = hdr_font, hdr_fill
    for i, w in enumerate([14, 12, 10, 14, 12, 12, 10, 34], 1):
        ws3.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path, len(data.get("targets", []))


if __name__ == "__main__":
    p, n = export_excel()
    print(f"✅ 导出 {n} 个检测目标 → {p}")
